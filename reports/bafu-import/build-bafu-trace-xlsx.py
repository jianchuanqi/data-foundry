#!/usr/bin/env python3
"""Build the consolidated BAFU import trace + progress workbook.

Sheets:
  1. 说明 Read me
  2. 导入进展 Summary
  3. 待人工校验 Human Review
  4. Process Trace      (every process in the universe, 11,747)
  5. Flow Trace         (every verified flow row, ~14,019)
  6. 转换映射 Conversion (effective source-flow -> canonical/mint decisions)
  7. Support Identities (contacts/sources/etc, ~14,179)
"""
import json, os, re, sys, glob
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Resolve paths relative to this script so the tool is committable and relocatable.
# Lives at <repo>/reports/bafu-import/build-bafu-trace-xlsx.py -> repo root is two dirs up.
# The run directory can be overridden with BAFU_RUN_DIR for a different import workspace.
HERE = os.path.dirname(os.path.abspath(__file__))
FOUNDRY = os.path.abspath(os.path.join(HERE, os.pardir, os.pardir))
RUN = os.environ.get(
    "BAFU_RUN_DIR",
    os.path.join(FOUNDRY, ".foundry/workspaces/bafu-full-import-20260607T080646Z"),
)
TIDAS_P = os.path.join(FOUNDRY, "inputs/BAFU-2025 Version 2 - TIDAS 2026-03-09/tidas/processes")
TIDAS_F = os.path.join(FOUNDRY, "inputs/BAFU-2025 Version 2 - TIDAS 2026-03-09/tidas/flows")
MERGED = os.path.join(RUN, "merged-verified-current")
OUT = os.path.join(HERE, "BAFU-导入trace与进展.xlsx")

# ---------------- helpers ----------------
def txt(x):
    if isinstance(x, list):
        for i in x:
            if isinstance(i, dict) and i.get("#text"):
                return i["#text"]
        return ""
    return x.get("#text", "") if isinstance(x, dict) else (str(x) if x is not None else "")

def loadl(p):
    out = []
    if os.path.exists(p):
        for ln in open(p):
            ln = ln.strip()
            if ln:
                try:
                    out.append(json.loads(ln))
                except Exception:
                    pass
    return out

def batch_of(report):
    m = re.search(r"batch-import-([^/]+)/", report or "")
    return m.group(1) if m else ""

def proc_name(pid):
    p = os.path.join(TIDAS_P, pid + ".json")
    if not os.path.exists(p):
        return ""
    try:
        d = json.load(open(p)); fl = d.get("processDataSet", d)
        return txt(fl["processInformation"]["dataSetInformation"]["name"]["baseName"])
    except Exception:
        return ""

def flow_name(fid):
    p = os.path.join(TIDAS_F, fid + ".json")
    if not os.path.exists(p):
        return ""
    try:
        d = json.load(open(p)); fl = d.get("flowDataSet", d)
        return txt(fl["flowInformation"]["dataSetInformation"]["name"]["baseName"])
    except Exception:
        return ""

# ---------------- load data ----------------
print("loading universe + ledgers ...")
universe = loadl(os.path.join(RUN, "coverage-v19/bafu-process-universe.coverage.jsonl"))
procs = loadl(os.path.join(MERGED, "ok.processes.verified.jsonl"))
flows = loadl(os.path.join(MERGED, "ok.flows.verified.jsonl"))
support = loadl(os.path.join(MERGED, "verified-support-identities.jsonl"))

# per-process import meta from the verified ledger (batch + timestamp)
pmeta = {}
for r in procs:
    pmeta[r["process_id"]] = (batch_of(r.get("report")), (r.get("generated_at_utc") or "")[:19])

# conversion decisions: union all identity-decisions, keep highest-priority per source
PRIO = {"v19-user": 100, "v17-user": 95, "v16-tierA": 90, "v20-mydata-reuse": 80,
        "v15-fullpool-leaf": 70, "v14-rejudge436-leaf": 60, "v13-b1-remap-leaf": 55,
        "v12-elementary-multi-leaf": 50, "v11-direct-process-leaf": 45}
def setprio(path):
    m = re.search(r"/decisions-([^/]+)/", path)
    return PRIO.get(m.group(1), 10) if m else 10
conv = {}   # source_entity_key -> (prio, row, set_name)
for dp in glob.glob(os.path.join(RUN, "decisions-*/identity-decisions.jsonl")):
    pr = setprio(dp)
    setname = re.search(r"/decisions-([^/]+)/", dp).group(1)
    for r in loadl(dp):
        key = r.get("source_entity_key") or r.get("source_dataset_id")
        if not key:
            continue
        if key not in conv or pr > conv[key][0]:
            conv[key] = (pr, r, setname)
print(f"  universe={len(universe)} procs={len(procs)} flows={len(flows)} support={len(support)} conv-decisions={len(conv)}")

print("reading process names (11,747 tidas files) ...")
pname = {u["process_id"]: proc_name(u["process_id"]) for u in universe}

# ---------------- styles ----------------
HDR = PatternFill("solid", fgColor="1F4E78"); HDRF = Font(name="Arial", bold=True, color="FFFFFF", size=11)
VERI = PatternFill("solid", fgColor="E2EFDA"); PEND = PatternFill("solid", fgColor="FCE4D6")
DEC = PatternFill("solid", fgColor="FFF2CC"); SUBH = PatternFill("solid", fgColor="D9E1F2")
F = Font(name="Arial", size=10); FB = Font(name="Arial", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top"); TOP = Alignment(vertical="top")
thin = Side(style="thin", color="D0D0D0"); BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

def header(ws, cols, widths, row=1):
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR; cell.font = HDRF; cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = None  # set later after data

wb = Workbook()

# ================= Sheet 1: 说明 =================
s = wb.active; s.title = "说明 Read me"
rows = [
 ("BAFU 2025 V2 导入 — Trace 与进展总表", True, 15),
 ("", False, 10),
 ("生成时间：2026-06-22 | 账户 dab05739-1a42-421b-8170-3b77146d1d64 | state_code=0", False, 10),
 ("数据来源：.foundry/workspaces/bafu-full-import-20260607T080646Z（coverage-v19 + merged-verified-current + 全部 decisions-* + tidas 源）", False, 10),
 ("", False, 10),
 ("导入现状：universe 11,747 / verified 11,740 / pending 7（gap 7，已达天花板）。", True, 11),
 ("", False, 10),
 ("各 sheet 说明：", True, 11),
 ("  • 导入进展 Summary —— 总体数字、按导入批次(round)的分布、关键里程碑。", False, 10),
 ("  • 待人工校验 Human Review —— 需要人工/BAFU 判断的项：7 个残留 scope、本会话已应用的语义映射(含对抗复核结论)、暂记残留的 TiO₂/Ulexite。", False, 10),
 ("  • Process Trace —— 每个 process 一行(11,747)：id/版本/名称/状态/导入批次/导入时间。", False, 10),
 ("  • Flow Trace —— 每条已核验 flow 一行(~14,019)：flow id/版本/所属 process/状态/批次/时间。", False, 10),
 ("  • 转换映射 Conversion —— 每个源 flow 的有效解析决策：reuse(复用既有 canonical) 或 mint(新建 My-Data)，含依据/置信度。这是导入转换的核心 lineage。", False, 10),
 ("  • Support Identities —— 导入过程创建/核验的支撑数据集(~14,179)：contact/source/flow property 等。", False, 10),
 ("", False, 10),
 ("注：源经 TIDAS 转换后 elementary flow 的 compartment 统一变成占位 'Emissions to air, unspecified'，语义以源名称为准（见 Human Review）。", False, 9),
]
for i, (t, b, sz) in enumerate(rows, 1):
    c = s.cell(row=i, column=1, value=t); c.font = Font(name="Arial", bold=b, size=sz); c.alignment = WRAP
s.column_dimensions["A"].width = 130

# ================= Sheet 2: 导入进展 Summary =================
s = wb.create_sheet("导入进展 Summary")
ver = sum(1 for u in universe if u.get("coverage_status") == "verified")
pend = sum(1 for u in universe if u.get("coverage_status") != "verified")
kv = [
 ("指标", "值"),
 ("Process universe (总过程数)", len(universe)),
 ("Verified (已入库已核验)", ver),
 ("Pending (待处理残留)", pend),
 ("Coverage gap", len(universe) - ver),
 ("Coverage %", f"{ver/len(universe)*100:.3f}%"),
 ("Flow rows verified (已核验 flow 行)", len(flows)),
 ("Support identities verified (支撑数据集)", len(support)),
 ("Conversion decisions (有效源→canonical 决策)", len(conv)),
]
for r in kv:
    s.append(list(r))
for c in range(1, 3):
    s.cell(row=1, column=c).fill = HDR; s.cell(row=1, column=c).font = HDRF
for r in range(2, len(kv) + 1):
    s.cell(row=r, column=1).font = FB
s.column_dimensions["A"].width = 44; s.column_dimensions["B"].width = 22
# batch distribution
s.append([]); s.append(["导入批次分布 (verified processes by import round)", ""])
s.cell(row=s.max_row, column=1).font = FB; s.cell(row=s.max_row, column=1).fill = SUBH
s.append(["批次 round", "process 数"])
for c in range(1, 3):
    s.cell(row=s.max_row, column=c).fill = HDR; s.cell(row=s.max_row, column=c).font = HDRF
bc = Counter(pmeta[p][0] for p in pmeta)
for k, v in bc.most_common():
    s.append([k or "(unknown)", v])

# ================= Sheet 3: 待人工校验 Human Review =================
s = wb.create_sheet("待人工校验 Human Review")
s.append(["类别", "对象 / process_id 或 flow", "名称", "状态 / 结论", "说明 / 需人工判断的点"])
for c in range(1, 6):
    s.cell(row=1, column=c).fill = HDR; s.cell(row=1, column=c).font = HDRF
s.row_dimensions[1].height = 28; s.freeze_panes = "A2"
hr = []
# 7 residual scopes
residual = [
 ("015cd7db-8a64-3f0b-ba15-6ff81e3d1eb4", "Polyvinylchloride, emulsion polymerised, at plant {RER}"),
 ("09cda8b1-b652-39b2-8761-c93aecc29f70", "Polypropylene, granulate, at plant {RER}"),
 ("48985f3e-97b4-3128-a1a8-a5ba91a1cb88", "Polyvinylchloride, suspension polymerised, at plant {RER}"),
 ("6738f024-eb64-3ef8-837e-05b84c6e43da", "Polyethylene, HDPE, granulate, at plant {RER}"),
 ("99a93e93-a8a8-34f3-88f3-6af0c21c4762", "Vinyl chloride, at plant {RER}"),
 ("c4dbbb99-22d0-37c1-8cc7-fc5a8c64329b", "Polyethylene, LDPE, granulate, at plant {RER}"),
 ("df878822-0deb-3a59-a6b4-396d94a3b654", "Polyethylene, LLDPE, granulate, at plant {RER}"),
]
for pid, nm in residual:
    blk = "TiO₂-ore + Ulexite" + (" + Particulates(PM10)" if "Vinyl chloride" not in nm else "")
    hr.append(["① 残留 scope (待BAFU讨论)", pid, nm, "PENDING",
               f"卡点 flow: {blk}。库里无 TiO₂矿/Ulexite 的质量等价 canonical（元素 Ti 占 TiO₂ 59.9%、元素硼占 ulexite ~13%）。需 BAFU 决定：新建 '…, in ground' 资源 flow / 按元素计 / 改源数据。"])
# applied semantic decisions (v17 6 + v19 3) with adversarial verdict
applied = [
 ("d614c930-c95b-5d2c-b894-ca6e46ce9ffd", "Methane, monochloro-, R-40", "→ chloromethane (08a91e70…9677)", "对抗复核 CONFIRM。R-40=chloromethane CAS 74-87-3。"),
 ("630ba118-3e3e-51f7-882e-35291d24b116", "Particulates, < 10 um (stationary)", "→ particles (PM10) (08a91e70…91be)", "对抗复核 CONFIRM。<10um 即 PM10。"),
 ("7cca9508-522a-54cd-9de0-a25e74b58a1b", "Energy, gross calorific value, in biomass", "→ …in biomass, primary forest (1de8efff)", "对抗复核 CONFIRM。同生物质热值能量资源(跨compartment)。"),
 ("f2ea6beb-946f-50b1-bb94-7f108d85a3dc", "Water, well, RER", "→ water (419682fe)", "对抗复核 CONFIRM。同 H₂O 水资源(跨compartment)。"),
 ("f8e2e9a9-5708-5cfe-9d5a-c5d770502307", "Calcium II", "→ calcium (08a91e70…9233)", "对抗复核 CONFIRM。Ca²⁺即钙(跨compartment)。"),
 ("927e15c3-55ff-540c-aeb4-d2ff56cb84b1", "Carbon dioxide, in air", "→ carbon dioxide (biogenic) (da174fac)", "对抗复核 CONFIRM。'in air'=生物吸收，biogenic 正确，无 fossil 歧义。"),
 ("f6f5acdd-a3be-57c1-881e-5431e8dcad6e", "Waste water", "→ Wastewater (d2d44ce1, Waste-flow型)", "研究纠正(你填的id解析不到)+对抗复核 ACCEPT。单scope试跑已验证可入库。"),
 ("f17a9daf-cd0e-5a27-99d7-3932565935ed", "Acidity, unspecified", "→ acid (as H+) (29066274)", "研究纠正(你记'新建'但库里已有)+对抗复核 ACCEPT。"),
]
for fid, nm, can, note in applied:
    hr.append(["② 已应用语义映射(本会话)", fid, nm, "APPLIED ✓", f"{can}。{note}"])
# recorded-but-pending
hr.append(["③ 暂记残留 / 待激活", "f78b5218 (TiO₂ in ilmenite)", "TiO2, 54% in ilmenite, 2.6% in crude ore", "RESIDUE", "你选'记残留待BAFU讨论'。无质量等价 canonical。"])
hr.append(["③ 暂记残留 / 待激活", "f7afa7fa (Ulexite)", "Ulexite", "RESIDUE", "你选'记残留待BAFU讨论'。库里无硼酸盐矿资源 flow。"])
hr.append(["③ 暂记残留 / 待激活", "ca266686 (Particulates)", "Particulates (泛指无粒径)", "RECORDED→PM10", "你选 PM10(decisions-v19 已写)，但其 scope 同时被 TiO₂/Ulexite 卡住，待那两者解决后自动放行。"])
for row in hr:
    s.append(row)
for r in range(2, s.max_row + 1):
    cat = s.cell(row=r, column=1).value
    for c in range(1, 6):
        cell = s.cell(row=r, column=c); cell.font = F; cell.alignment = WRAP; cell.border = BORD
    st = s.cell(row=r, column=4).value or ""
    s.cell(row=r, column=4).fill = PEND if ("PENDING" in st or "RESIDUE" in st) else (VERI if "✓" in st else DEC)
for i, w in enumerate([26, 40, 40, 30, 70], 1):
    s.column_dimensions[get_column_letter(i)].width = w

# ================= Sheet 4: Process Trace =================
s = wb.create_sheet("Process Trace")
header(s, ["process_id", "version", "名称 baseName", "状态", "导入批次 round", "导入时间 UTC"],
       [40, 10, 60, 12, 22, 22])
for u in universe:
    pid = u["process_id"]; st = u.get("coverage_status", "")
    b, ts = pmeta.get(pid, ("", ""))
    s.append([pid, u.get("process_version", ""), pname.get(pid, ""), st, b, ts])
s.auto_filter.ref = f"A1:F{s.max_row}"
# light status banding (only the status column, fast)
for r in range(2, s.max_row + 1):
    stv = s.cell(row=r, column=4).value
    s.cell(row=r, column=4).fill = VERI if stv == "verified" else PEND

# ================= Sheet 5: Flow Trace =================
s = wb.create_sheet("Flow Trace")
header(s, ["flow_id", "version", "所属 process_id", "状态", "导入批次 round", "导入时间 UTC"],
       [40, 12, 40, 12, 22, 22])
for r in flows:
    s.append([r.get("dataset_id", ""), r.get("dataset_version", ""), r.get("process_id", ""),
              r.get("status", ""), batch_of(r.get("report")), (r.get("generated_at_utc") or "")[:19]])
s.auto_filter.ref = f"A1:F{s.max_row}"

# ================= Sheet 6: 转换映射 Conversion =================
s = wb.create_sheet("转换映射 Conversion")
header(s, ["源 flow id (BAFU)", "源 flow 名称", "决策", "→ canonical/mint id", "canonical 名称",
           "决策来源 set", "置信度", "依据 basis"],
       [40, 34, 22, 40, 30, 22, 10, 60])
# read source flow names for the conversion rows
conv_rows = sorted(conv.values(), key=lambda x: -x[0])
print(f"reading {len(conv_rows)} source-flow names for conversion sheet ...")
for prio, r, setname in conv_rows:
    src = r.get("source_dataset_id", "")
    dec = r.get("identity_decision") or r.get("decision") or ""
    can = r.get("canonical") or {}
    canid = can.get("ref_object_id") or r.get("canonical_flow_id") or ""
    canname = can.get("short_description") or r.get("canonical_short_description") or ""
    s.append([src, flow_name(src), dec, canid, canname, setname,
              r.get("confidence", ""), (r.get("basis") or "")[:300]])
s.auto_filter.ref = f"A1:H{s.max_row}"

# ================= Sheet 7: Support Identities =================
s = wb.create_sheet("Support Identities")
header(s, ["identity_key", "dataset_type", "dataset_id", "version", "状态", "来源 source"],
       [50, 16, 40, 12, 12, 30])
for r in support:
    s.append([r.get("identity_key", ""), r.get("dataset_type", ""), r.get("dataset_id", ""),
              r.get("dataset_version", ""), r.get("status", ""), r.get("source", "")])
s.auto_filter.ref = f"A1:F{s.max_row}"

# ---------------- save ----------------
print("saving workbook ...")
wb.save(OUT)
print("SAVED:", OUT)
print("SHEET ROWS:", {ws.title: ws.max_row for ws in wb.worksheets})
# emit counts for verification
print("VERIFY_COUNTS", json.dumps({
    "universe": len(universe), "verified": ver, "pending": pend,
    "flows": len(flows), "support": len(support), "conversion_decisions": len(conv),
}))
